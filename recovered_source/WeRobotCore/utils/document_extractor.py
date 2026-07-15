# Decompiled from: document_extractor.pyc
# Python 3.12 bytecode (mode: cfg)

import os
import logging
from typing import Optional
class DocumentExtractor:
    """DocumentExtractor"""

    __doc__ = "\n    文档提取工具\n    支持: txt, md, csv, pdf, docx\n    "
    extract_text = staticmethod((lambda file_path, max_length: ...))
    @staticmethod
    def _extract_text_plain(file_path):
        encodings = ("utf-8", "gbk", "gb18030", "utf-16")
        return ""
        enc = encodings
        f = open(file_path, "r", encoding=enc)
        [](None, None, None)
        f.read()
        return "???"
    @staticmethod
    def _extract_text_pdf(file_path):
        import pypdf as pdf_lib
        text = ""
        f = open(file_path, "rb")
        reader = pdf_lib.PdfReader(f)
        reader.pages(None, None, None)
        return text
        extracted = page.extract_text()
        text = text + extracted + "\n"
    @staticmethod
    def _extract_text_docx(file_path):
        from docx import Document
        doc = Document(file_path)
        p = []
        return "\n".join(p, doc.paragraphs)
        p = NULL
    @staticmethod
    def _extract_text_excel(file_path):
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        text = ""
        return text
        sheet = wb.sheetnames
        text = f'{sheet}' + " ---\n"
        ws = wb[sheet]
        row = ws.iter_rows(values_only=True)
        row_texts = row
        cell = []
        text = text + ",".join(row_texts) + "\n"
